from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, session, jsonify, make_response
)
from functools import wraps
from modules.recipient import add_recipient, get_recipients_by_type
from modules.announcement import save_announcement, save_log
from database.mongo import get_db
from modules.user import (
    get_all_users, get_user_by_id, add_user,
    update_user, toggle_user_status, bulk_add_users,
    get_users_by_target
)
from modules.audit import log_audit
from utils.email_service import send_email
from modules.scheduler import init_scheduler
from modules.security import (
    init_security, check_password, get_all_system_users,
    add_system_user, update_system_user, delete_system_user
)

import config
import os
import csv
import io
import openpyxl
import datetime

app = Flask(__name__)
app.secret_key = os.urandom(24)

# Start background scheduler
init_scheduler()

# Securely bootstrap admin credentials to DB if empty
with app.app_context():
    init_security()

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def login_required(f):
    """Redirect to /login if the user is not authenticated."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_email" not in session:
            flash("Please log in to access that page.", "warning")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def super_admin_required(f):
    """Allow ONLY super_admin."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_role") != "super_admin":
            flash("Requires Super Admin privileges.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Allow super_admin and admin (Recipient management)."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_role") not in ("super_admin", "admin"):
            flash("You do not have permission to access that page.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated
    
def operator_required(f):
    """Allow super_admin, admin, and operator."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if session.get("user_role") not in ("super_admin", "admin", "operator"):
            flash("You do not have permission to access that page.", "danger")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route("/login", methods=["GET", "POST"])
def login():
    if "user_email" in session:
        return redirect(url_for("dashboard"))
        
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        
        db = get_db()
        # Requirement: Email must exist in users collection 
        # (Assuming system_users for this protected dashboard context)
        user = db.system_users.find_one({"email": email, "status": "active"})
        
        if user and check_password(password, user.get("password_hash")):
            # Set specific session variables requested
            session["user_id"] = str(user["_id"])
            session["user_email"] = user.get("email")
            session["user_role"] = user.get("role", "operator")
            session["user_name"] = user.get("name", user.get("username"))
            
            # Legacy/Compatibility keys if needed elsewhere (keeping role for navbar)
            session["username"] = user.get("username")
            session["role"] = user.get("role")
            
            log_audit(user_id=session["user_email"], action="User login", description=f"Successful login for {session['user_name']} ({session['user_role']})", ip_address=request.remote_addr)
            flash(f"Welcome, {session['user_name']}! You are logged in as {session['user_role']}.", "success")
            return redirect(url_for("dashboard"))
            
        flash("Invalid email or password.", "danger")
        
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have been logged out.", "info")
    return redirect(url_for("login"))


# ---------------------------------------------------------------------------
# Core announcement routes (unchanged)
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    if "user_email" in session:
        return redirect(url_for("dashboard"))
    
    db = get_db()
    
    # --- 1. Public Statistics ---
    total_users = db.users.count_documents({"status": "active"})
    
    # New users today (UTC)
    today_start = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    new_users_today = db.users.count_documents({
        "status": "active",
        "created_at": {"$gte": today_start}
    })
    
    # --- 2. Public Activity (Audit Logs) ---
    # We show generic actions like 'Creation', 'Update' but hide sensitive IDs if necessary
    recent_activities = list(db.audit_logs.find().sort("timestamp", -1).limit(5))
    
    # --- 3. Latest Announcements ---
    public_announcements = list(db.announcements.find({"status": "Sent"}).sort("created_at", -1).limit(5))
    
    stats = {
        "total_users": total_users,
        "new_users_today": new_users_today
    }
    
    return render_template("landing.html", 
                           announcements=public_announcements,
                           stats=stats,
                           activities=recent_activities)

@app.route("/dashboard")
@login_required
@operator_required
def dashboard():
    db = get_db()
    
    # --- 1. Top Level Statistics ---
    total_users = db.users.count_documents({"status": "active"})
    total_students = db.users.count_documents({"status": "active", "type": "student"})
    total_faculty = db.users.count_documents({"status": "active", "type": "faculty"})
    total_announcements = db.announcements.count_documents({})
    
    total_reads = db.announcement_reads.count_documents({"read_status": True})
    total_dispatches = db.announcement_reads.count_documents({})
    global_read_rate = round((total_reads / total_dispatches) * 100) if total_dispatches > 0 else 0
    
    stats = {
        "total_users": total_users,
        "total_students": total_students,
        "total_faculty": total_faculty,
        "total_announcements": total_announcements,
        "global_read_rate": global_read_rate
    }
    
    # --- 2. Chart.js Data (Last 7 Days) ---
    today = datetime.datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    dates = [(today - datetime.timedelta(days=i)) for i in range(6, -1, -1)]
    chart_labels = [d.strftime("%b %d") for d in dates]
    
    announcement_activity_data = []
    user_growth_data = []
    
    for d in dates:
        next_d = d + datetime.timedelta(days=1)
        a_count = db.announcements.count_documents({"created_at": {"$gte": d, "$lt": next_d}})
        u_count = db.users.count_documents({"created_at": {"$gte": d, "$lt": next_d}})
        announcement_activity_data.append(a_count)
        user_growth_data.append(u_count)
        
    charts = {
        "labels": chart_labels,
        "announcements": announcement_activity_data,
        "users": user_growth_data
    }
    
    # --- 3. Recent Activity Logs (Using Audit Logs for high-level events) ---
    recent_activity = list(db.audit_logs.find().sort("timestamp", -1).limit(6))
    
    # --- 4. Recent Announcements list ---
    recent_announcements = list(db.announcements.find().sort("created_at", -1).limit(5))
    
    # Attach read statistics
    for ann in recent_announcements:
        ann_id = ann.get("announcement_id", str(ann.get("_id")))
        t_recipients = db.announcement_reads.count_documents({"announcement_id": ann_id})
        r_count = db.announcement_reads.count_documents({"announcement_id": ann_id, "read_status": True})
        
        ann["stat_total"] = t_recipients
        ann["stat_read"] = r_count
        ann["stat_unread"] = t_recipients - r_count
        ann["stat_percent"] = round((r_count / (t_recipients or 1)) * 100)
        
    return render_template("index.html", 
                           announcements=recent_announcements,
                           stats=stats,
                           charts=charts,
                           recent_activity=recent_activity)

# --- Email Tracking Route --- 
@app.route("/track/<announcement_id>/<email>")
def track_read(announcement_id, email):
    db = get_db()
    
    # Mark as read
    db.announcement_reads.update_one(
        {"announcement_id": announcement_id, "user_email": email},
        {"$set": {
            "read_status": True, 
            "read_time": datetime.datetime.utcnow()
        }},
        upsert=True
    )
    
    # Return 1x1 transparent GIF
    transparent_gif = b'GIF89a\x01\x00\x01\x00\x80\x00\x00\x00\x00\x00\x00\x00\x00!\xf9\x04\x01\x00\x00\x00\x00,\x00\x00\x00\x00\x01\x00\x01\x00\x00\x02\x02D\x01\x00;'
    response = make_response(transparent_gif)
    response.headers.set('Content-Type', 'image/gif')
    response.headers.set('Cache-Control', 'no-cache, no-store, must-revalidate')
    return response


@app.route("/add-email", methods=["GET", "POST"])
@login_required
@admin_required
def add_email():
    if request.method == "POST":
        name = request.form.get("name")
        email = request.form.get("email")
        r_type = request.form.get("type")
        course = request.form.get("course")
        add_recipient(name, email, r_type, course)
        log_audit(user_id=session.get("username", "anonymous"), action="User creation", description=f"Admin created recipient: {name}", ip_address=request.remote_addr)
        flash("Email added successfully!", "success")
        return redirect(url_for("add_email"))
    return render_template("add_email.html")


@app.route("/preview", methods=["POST"])
@login_required
@operator_required
def preview():
    title = request.form.get("title", "").strip()
    content = request.form.get("content", "").strip()
    priority = request.form.get("priority", "normal")
    
    target_type = request.form.get("target_type")
    target_department = request.form.get("department", "").strip()
    target_course = request.form.get("course", "").strip()
    target_year = request.form.get("year", "").strip()
    
    # --- Integration for Scheduled feature ---
    # Convert local HTML datetime-local string to UTC datetime object if provided
    schedule_time_str = request.form.get("schedule_time", "").strip()
    schedule_time = None
    if schedule_time_str:
        try:
            # HTML5 datetime-local format is 'YYYY-MM-DDTHH:MM'
            local_dt = datetime.datetime.strptime(schedule_time_str, "%Y-%m-%dT%H:%M")
            # For simplicity without pytz, we'll store the exact datetime structure. 
            # Note: Production scale often requires rigorous tz conversion.
            schedule_time = local_dt
        except ValueError:
            pass
    
    if not title or not content or not target_type:
        flash("Please fill all required fields.", "danger")
        return redirect(url_for("index"))
        
    target_dict = {
        "target_type": target_type,
        "department": target_department,
        "course": target_course,
        "year": target_year
    }
    
    # Optional: fetch a count to show on preview page
    matching_users = get_users_by_target(target_dict)
    
    session["announcement_data"] = {
        "title": title,
        "content": content,
        "priority": priority,
        "target_dict": target_dict,
        # Store as ISO format string for JSON session serialization
        "schedule_time": schedule_time.isoformat() if schedule_time else None
    }
    
    return render_template("preview.html", 
        title=title, 
        content=content, 
        priority=priority,
        target_dict=target_dict,
        schedule_time=schedule_time_str, # Passed back for display
        match_count=len(matching_users)
    )


@app.route("/send-announcement", methods=["POST"])
@login_required
@operator_required
def send_announcement():
    data = session.get("announcement_data")
    if not data:
        flash("No announcement data found. Please try again.", "danger")
        return redirect(url_for("index"))
        
    title = data["title"]
    content = data["content"]
    priority = data.get("priority", "normal")
    target_dict = data.get("target_dict", {})
    
    schedule_time_iso = data.get("schedule_time")
    schedule_time = datetime.datetime.fromisoformat(schedule_time_iso) if schedule_time_iso else None
    
    # If scheduled in future, status is Scheduled. Else Sent.
    # Note: Using exact DB-time alignment rules (utc) or relying on local time.
    # We will let the scheduler resolve it properly based on the dt object.
    status = "Scheduled" if schedule_time else "Sent"
    
    created_by = session.get("username", "anonymous")
    
    announcement_id = save_announcement(
        title=title, 
        content=content, 
        priority=priority, 
        target_type=target_dict.get('target_type'), 
        target_department=target_dict.get('department'), 
        target_course=target_dict.get('course'), 
        target_year=target_dict.get('year'), 
        created_by=created_by,
        status=status,
        schedule_time=schedule_time
    )
    
    if status == "Sent":
        recipients = get_users_by_target(target_dict)
        success_count = 0
        fail_count = 0
        db = get_db()
        
        for r in recipients:
            r_status = "Failed"
            
            # Initialize Read Tracking
            db.announcement_reads.insert_one({
                "announcement_id": announcement_id,
                "user_email": r["email"],
                "read_status": False,
                "read_time": None
            })
            
            # send_email now takes (to_email, title, content, announcement_id)
            if send_email(r["email"], title, content, announcement_id):
                r_status = "Sent"
                success_count += 1
            else:
                fail_count += 1
            save_log(announcement_id, r["name"], r["email"], r["type"], r_status)
            
        session.pop("announcement_data", None)
        log_audit(user_id=created_by, action="Announcement creation", description=f"Admin created announcement: {title}", ip_address=request.remote_addr)
        flash(f"Announcement sent! Success: {success_count}, Failed: {fail_count}", "success")
    else:
        # Scheduled
        session.pop("announcement_data", None)
        log_audit(user_id=created_by, action="Announcement creation", description=f"Admin scheduled announcement: {title}", ip_address=request.remote_addr)
        flash(f"Announcement scheduled for {schedule_time.strftime('%Y-%m-%d %H:%M')}. It will be sent automatically.", "info")
        
    return redirect(url_for("index"))

@app.route("/audit-logs")
@login_required
@super_admin_required
def audit_logs():
    db = get_db()
    logs = list(db.audit_logs.find().sort("timestamp", -1).limit(200))
    return render_template("audit_logs.html", logs=logs)

@app.route("/admin/staff")
@login_required
@super_admin_required
def manage_staff():
    staff = get_all_system_users()
    smtp_config = {
        "email": config.SMTP_EMAIL or "",
        "host": os.getenv("SMTP_HOST", "smtp.gmail.com"),
        "port": os.getenv("SMTP_PORT", "587"),
    }
    return render_template("manage_staff.html", staff=staff, smtp_config=smtp_config)

@app.route("/admin/staff/add", methods=["POST"])
@login_required
@super_admin_required
def add_staff_route():
    username = request.form.get("username")
    name = request.form.get("name")
    email = request.form.get("email")
    password = request.form.get("password")
    role = request.form.get("role", "operator")

    if not all([username, name, email, password]):
        flash("All fields are required.", "danger")
        return redirect(url_for("manage_staff"))

    success, err = add_system_user(username, email, name, password, role)
    if success:
        log_audit(user_id=session.get("user_email"), action="Staff creation", description=f"Super Admin created {role}: {name}", ip_address=request.remote_addr)
        flash(f"Staff member '{name}' added successfully!", "success")
    else:
        flash(f"Error: {err}", "danger")
    return redirect(url_for("manage_staff"))

@app.route("/admin/staff/edit/<user_id>", methods=["POST"])
@login_required
@super_admin_required
def edit_staff_route(user_id):
    name = request.form.get("name")
    email = request.form.get("email")
    role = request.form.get("role")

    success, err = update_system_user(user_id, name, email, role)
    if success:
        log_audit(user_id=session.get("user_email"), action="Staff update", description=f"Super Admin updated staff: {name}", ip_address=request.remote_addr)
        flash("Staff details updated successfully!", "success")
    else:
        flash(f"Error: {err}", "danger")
    return redirect(url_for("manage_staff"))

@app.route("/admin/staff/delete/<user_id>", methods=["POST"])
@login_required
@super_admin_required
def delete_staff_route(user_id):
    # Protection against self-deletion
    if user_id == session.get("user_id"):
        flash("You cannot delete your own account.", "warning")
        return redirect(url_for("manage_staff"))

    success, err = delete_system_user(user_id)
    if success:
        log_audit(user_id=session.get("user_email"), action="Staff deletion", description=f"Super Admin deleted staff ID: {user_id}", ip_address=request.remote_addr)
        flash("Staff member removed.", "success")
    else:
        flash(f"Error: {err}", "danger")
    return redirect(url_for("manage_staff"))

@app.route("/admin/settings/smtp", methods=["POST"])
@login_required
@super_admin_required
def update_smtp_settings():
    """Allow Super Admin to update SMTP config from the UI."""
    smtp_email = request.form.get("smtp_email", "").strip()
    smtp_password = request.form.get("smtp_password", "").strip()
    smtp_host = request.form.get("smtp_host", "smtp.gmail.com").strip()
    smtp_port = request.form.get("smtp_port", "587").strip()

    if not smtp_email:
        flash("SMTP Email cannot be empty.", "danger")
        return redirect(url_for("manage_staff"))

    # Update .env file
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    try:
        with open(env_path, "r") as f:
            lines = f.readlines()

        new_lines = []
        keys_found = set()
        for line in lines:
            if line.startswith("SMTP_EMAIL="):
                new_lines.append(f"SMTP_EMAIL={smtp_email}\n")
                keys_found.add("SMTP_EMAIL")
            elif line.startswith("SMTP_PASSWORD=") and smtp_password:
                new_lines.append(f"SMTP_PASSWORD={smtp_password}\n")
                keys_found.add("SMTP_PASSWORD")
            elif line.startswith("SMTP_HOST="):
                new_lines.append(f"SMTP_HOST={smtp_host}\n")
                keys_found.add("SMTP_HOST")
            elif line.startswith("SMTP_PORT="):
                new_lines.append(f"SMTP_PORT={smtp_port}\n")
                keys_found.add("SMTP_PORT")
            else:
                new_lines.append(line)

        # Add any keys that weren't in the file
        if "SMTP_EMAIL" not in keys_found:
            new_lines.append(f"SMTP_EMAIL={smtp_email}\n")
        if "SMTP_PASSWORD" not in keys_found and smtp_password:
            new_lines.append(f"SMTP_PASSWORD={smtp_password}\n")
        if "SMTP_HOST" not in keys_found:
            new_lines.append(f"SMTP_HOST={smtp_host}\n")
        if "SMTP_PORT" not in keys_found:
            new_lines.append(f"SMTP_PORT={smtp_port}\n")

        with open(env_path, "w") as f:
            f.writelines(new_lines)

        # Update the live config module attributes immediately (no restart needed)
        import config as _cfg
        _cfg.SMTP_EMAIL = smtp_email
        if smtp_password:
            _cfg.SMTP_PASSWORD = smtp_password
        os.environ["SMTP_EMAIL"] = smtp_email
        if smtp_password:
            os.environ["SMTP_PASSWORD"] = smtp_password

        log_audit(
            user_id=session.get("user_email"),
            action="SMTP configuration update",
            description=f"Super Admin updated SMTP settings: email={smtp_email}, host={smtp_host}, port={smtp_port}",
            ip_address=request.remote_addr
        )
        flash("✅ SMTP settings updated successfully! Email service is now using the new configuration.", "success")
    except Exception as e:
        flash(f"Error updating SMTP settings: {e}", "danger")

    return redirect(url_for("manage_staff"))

@app.route("/delete-announcement/<announcement_id>", methods=["POST"])
@login_required
@admin_required
def delete_announcement(announcement_id):
    db = get_db()
    ann = db.announcements.find_one({"announcement_id": announcement_id})
    if ann:
        title = ann.get('title', 'Unknown')
        db.announcements.delete_one({"announcement_id": announcement_id})
        # Note: Depending on rules, you might also cascade delete reads/logs.
        log_audit(user_id=session.get("username"), action="Announcement deletion", description=f"Admin deleted announcement: {title}", ip_address=request.remote_addr)
        flash("Announcement deleted successfully.", "success")
    else:
        flash("Announcement not found.", "danger")
    return redirect(url_for("index"))


# ---------------------------------------------------------------------------
# User Management routes (admin / main_user only)
# ---------------------------------------------------------------------------

@app.route("/users")
@login_required
@admin_required
def users():
    all_users = get_all_users()
    return render_template("users.html", users=all_users)


@app.route("/users/add", methods=["POST"])
@login_required
@admin_required
def add_user_route():
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()
    user_type = request.form.get("type", "student").strip()
    department = request.form.get("department", "").strip()
    course = request.form.get("course", "").strip()
    year = request.form.get("year", "").strip()
    created_by = session.get("username")

    if not name or not email:
        flash("Name and Email are required.", "danger")
        return redirect(url_for("users"))

    uid, err = add_user(name, email, user_type, department, course, year, created_by)
    if err:
        flash(f"Error: {err}", "danger")
    else:
        flash(f"User '{name}' added successfully!", "success")
    return redirect(url_for("users"))


@app.route("/users/edit/<user_id>", methods=["POST"])
@login_required
@admin_required
def edit_user_route(user_id):
    name = request.form.get("name", "").strip()
    email = request.form.get("email", "").strip()
    user_type = request.form.get("type", "student").strip()
    department = request.form.get("department", "").strip()
    course = request.form.get("course", "").strip()
    year = request.form.get("year", "").strip()

    ok, err = update_user(user_id, name, email, user_type, department, course, year)
    if ok:
        flash("User updated successfully!", "success")
    else:
        flash(f"Error updating user: {err}", "danger")
    return redirect(url_for("users"))


@app.route("/users/toggle/<user_id>", methods=["POST"])
@login_required
@admin_required
def toggle_user_route(user_id):
    ok, result = toggle_user_status(user_id)
    if ok:
        flash(f"User status changed to '{result}'.", "success")
    else:
        flash(f"Error: {result}", "danger")
    return redirect(url_for("users"))


@app.route("/users/bulk-upload", methods=["POST"])
@login_required
@admin_required
def bulk_upload():
    file = request.files.get("csv_file")  # Input name in form is still csv_file, or we change to upload_file
    if not file or file.filename == "":
        flash("No file selected.", "danger")
        return redirect(url_for("users"))
    
    filename_lower = file.filename.lower()
    mimetype = file.mimetype
    
    # MIME-type validation for enhanced security
    allowed_mimetypes = [
        'text/csv', 
        'application/vnd.ms-excel', 
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    ]
    
    if mimetype not in allowed_mimetypes and not (filename_lower.endswith(".csv") or filename_lower.endswith(".xlsx")):
        flash("Invalid file type. Only .csv and .xlsx files are accepted.", "danger")
        return redirect(url_for("users"))

    rows = []
    try:
        if filename_lower.endswith(".csv"):
            stream = io.StringIO(file.stream.read().decode("utf-8-sig"), newline=None)
            reader = csv.DictReader(stream)
            rows = list(reader)
        elif filename_lower.endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
            sheet = wb.active
            header = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Skip completely empty rows
                if any(row):
                    rows.append(dict(zip(header, row)))
    except Exception as e:
        flash(f"Error parsing file: {str(e)}", "danger")
        return redirect(url_for("users"))

    if not rows:
        flash("The uploaded file is empty or missing headers.", "warning")
        return redirect(url_for("users"))

    # Convert headers from xlsx to lowercase strings safely
    rows = [{str(k).lower().strip() if k else "": v for k, v in r.items()} for r in rows]

    created_by = session.get("username")
    total, success, duplicates, errors = bulk_add_users(rows, created_by)

    if success > 0:
        log_audit(user_id=created_by, action="Bulk user import", description=f"Admin uploaded {success} users via Excel", ip_address=request.remote_addr)

    summary_msg = f"Total rows: {total}<br>Imported: {success}<br>Duplicates skipped: {duplicates}<br>Errors: {len(errors)}"
    flash(summary_msg, "success" if success > 0 else "warning")
    
    # We can also flash the first few errors if user wants to see them
    for e in errors[:5]:
        flash(e, "danger")
    if len(errors) > 5:
        flash(f"...and {len(errors) - 5} more errors.", "danger")
        
    return redirect(url_for("users"))


@app.route("/users/download-template")
@login_required
@admin_required
def download_template():
    headers = ["name", "email", "type", "department", "course", "year"]
    sample = ["John Doe", "john@example.com", "student", "Computer Science", "BCA", "2"]
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    writer.writerow(sample)
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=users_template.csv"
    response.headers["Content-type"] = "text/csv"
    return response


@app.route("/users/download-excel-template")
@login_required
@admin_required
def download_excel_template():
    headers = ["Name", "Email", "Type", "Department", "Course", "Year"]
    sample1 = ["Rahul Sharma", "rahul@gmail.com", "Student", "CS", "BCA", "2"]
    sample2 = ["Anjali Verma", "anjali@gmail.com", "Faculty", "IT", "", ""]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(headers)
    ws.append(sample1)
    ws.append(sample2)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    response = make_response(out.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=users_template.xlsx"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


# ---------------------------------------------------------------------------
# Admin: Dedicated Add-User Page
# ---------------------------------------------------------------------------

@app.route("/admin/add-user", methods=["GET", "POST"])
@login_required
@super_admin_required
def admin_add_user():
    form_data = {}   # repopulate on validation error
    errors = {}

    if request.method == "POST":
        name       = request.form.get("name", "").strip()
        email      = request.form.get("email", "").strip()
        user_type  = request.form.get("type", "").strip()
        department = request.form.get("department", "").strip()
        course     = request.form.get("course", "").strip()
        year       = request.form.get("year", "").strip()
        status     = request.form.get("status", "active").strip()

        form_data = {
            "name": name, "email": email, "type": user_type,
            "department": department, "course": course,
            "year": year, "status": status,
        }

        # --- Server-side validation ---
        if not name:
            errors["name"] = "Full name is required."
        if not email:
            errors["email"] = "Email address is required."
        if not user_type:
            errors["type"] = "User type is required."

        if not errors:
            uid, err = add_user(
                name, email, user_type, department, course, year,
                created_by=session.get("username")
            )
            if err:
                errors["email"] = err   # duplicate email returns here
            else:
                # --- Audit log ---
                log_audit(
                    user_id=session.get("username"),
                    action="User creation",
                    description=f"Admin created user: {name} ({email})",
                    ip_address=request.remote_addr
                )
                # Override status if admin chose inactive
                if status == "inactive":
                    from modules.user import toggle_user_status as _toggle
                    from database.mongo import get_db as _get_db
                    from bson import ObjectId
                    _get_db().users.update_one(
                        {"_id": ObjectId(uid)}, {"$set": {"status": "inactive"}}
                    )
                flash(
                    f"User '{name}' ({email}) created successfully!",
                    "success"
                )
                return redirect(url_for("admin_add_user"))

    return render_template("admin_add_user.html", form=form_data, errors=errors)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
